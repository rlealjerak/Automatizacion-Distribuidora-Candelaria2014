"""
Parser tests. CSV/Excel are exact round-trips (easy to construct
precisely). PDF is tested against a real generated PDF table (via
reportlab, test-only dependency - see pyproject.toml) rather than asserted
against nothing, since CLAUDE.md flags PDF catalogs as the messiest input
this system has to handle and pdfplumber's table detection is the one
piece of this parser that can't be unit-tested by reasoning alone.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from adc_backend.db.core_models import SourceFileType
from adc_backend.modules.ingestion.parsers import (
    _rows_from_grid,
    detect_file_type,
    parse_csv,
    parse_excel,
    parse_pdf,
)


def test_detect_file_type():
    assert detect_file_type("catalog.xlsx") == SourceFileType.EXCEL
    assert detect_file_type("catalog.CSV") == SourceFileType.CSV
    assert detect_file_type("catalog.pdf") == SourceFileType.PDF
    with pytest.raises(ValueError):
        detect_file_type("catalog.docx")


def test_rows_from_grid_finds_header_and_dedupes_columns():
    grid = [
        [None, None],  # blank spacer row before the real header - common in supplier files
        ["ITEM #", "DESC", "DESC"],  # duplicate header cell
        ["ABC-1", "Widget", "Extra"],
        [None, None, None],  # blank row mid-data - should be skipped, not counted
        ["ABC-2", "Gadget", "Extra2"],
    ]
    warnings: list[str] = []
    headers, rows = _rows_from_grid(grid, warnings)
    assert headers == ["ITEM #", "DESC", "DESC_1"]
    assert len(rows) == 2
    assert rows[0].row_number == 1
    assert rows[0].raw_data == {"ITEM #": "ABC-1", "DESC": "Widget", "DESC_1": "Extra"}
    assert rows[1].row_number == 2  # blank row didn't consume a row_number


def test_rows_from_grid_empty_input_warns_not_crashes():
    warnings: list[str] = []
    headers, rows = _rows_from_grid([], warnings)
    assert headers == []
    assert rows == []
    assert warnings


def test_parse_csv_round_trip():
    content = b"ITEM #,UNIT PRICE,QTY\nABC-1,2.90,180\nABC-2,1.45,60\n"
    result = parse_csv(content)
    assert result.headers == ["ITEM #", "UNIT PRICE", "QTY"]
    assert len(result.rows) == 2
    assert result.rows[0].raw_data == {"ITEM #": "ABC-1", "UNIT PRICE": "2.90", "QTY": "180"}


def test_parse_csv_strips_bom():
    content = b"\xef\xbb\xbfITEM #,PRICE\nABC-1,2.90\n"
    result = parse_csv(content)
    assert result.headers[0] == "ITEM #"  # not "﻿ITEM #"


def test_parse_excel_round_trip():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SUPPLIER ITEM", "SALE PRICE"])
    ws.append(["XYZ-1", 4.5])
    ws.append(["XYZ-2", 7.25])
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_excel(buf.getvalue())
    assert result.headers == ["SUPPLIER ITEM", "SALE PRICE"]
    assert len(result.rows) == 2
    assert result.rows[0].raw_data == {"SUPPLIER ITEM": "XYZ-1", "SALE PRICE": "4.5"}


def test_parse_excel_multi_sheet_warns_and_uses_first():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ITEM", "PRICE"])
    ws.append(["A1", "1.00"])
    wb.create_sheet("Sheet2")
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_excel(buf.getvalue())
    assert len(result.rows) == 1
    assert any("2 sheets" in w for w in result.warnings)


def _build_test_pdf_with_table(header: list[str], data_rows: list[list[str]]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    table_data = [header, *data_rows]
    table = Table(table_data)
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    doc.build([table])
    return buf.getvalue()


def test_parse_pdf_extracts_real_table():
    pdf_bytes = _build_test_pdf_with_table(
        ["ITEM #", "UNIT PRICE", "CASE QTY"],
        [["ABC-1", "2.90", "180"], ["ABC-2", "1.45", "60"]],
    )
    result = parse_pdf(pdf_bytes)
    assert result.headers == ["ITEM #", "UNIT PRICE", "CASE QTY"]
    assert len(result.rows) == 2
    assert result.rows[0].raw_data == {"ITEM #": "ABC-1", "UNIT PRICE": "2.90", "CASE QTY": "180"}


def test_parse_pdf_no_tables_warns_not_crashes():
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph

    doc.build([Paragraph("Just marketing text, no table here.", getSampleStyleSheet()["Normal"])])

    result = parse_pdf(buf.getvalue())
    assert result.rows == []
    assert any("No tables detected" in w for w in result.warnings)

"""
Parsers that turn an uploaded file's raw bytes into rows of
{original_header: cell_value}, ready to become `RawLineItem.raw_data`.

Deliberately does NOT attempt normalization here (column renaming, type
coercion, unit parsing) - that's build-order step 4. This layer's only job
is "turn bytes into a table," as literally as possible, so raw_data always
reflects exactly what was in the source file.

Honest limitation, flagged rather than hidden: CLAUDE.md describes PDF
supplier catalogs as multi-column, image-heavy, with inconsistent headers
and noise columns - a fully robust general-purpose extractor for that is
not something to claim without real sample files to test against, which
aren't available in this repo. `parse_pdf` below uses pdfplumber's default
table detection with a first-non-empty-row-is-header heuristic, the same
one `parse_excel`/`parse_csv` use. Whatever it gets wrong on a given
supplier's PDF is exactly what step 4's confidence-scored mapping review
and the manual-review-queue routing exist to catch - this parser is not
meant to be the last line of defense against messy input.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

import openpyxl
import pdfplumber

from adc_backend.db.core_models import SourceFileType


@dataclass
class ParsedRow:
    row_number: int  # 1-based position among data rows (header excluded)
    raw_data: dict[str, str]


@dataclass
class ParseResult:
    headers: list[str]
    rows: list[ParsedRow]
    warnings: list[str] = field(default_factory=list)


def detect_file_type(filename: str) -> SourceFileType:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls", "xlsm"):
        return SourceFileType.EXCEL
    if ext == "csv":
        return SourceFileType.CSV
    if ext == "pdf":
        return SourceFileType.PDF
    raise ValueError(f"Unrecognized file extension for {filename!r} - expected .xlsx/.xls, .csv, or .pdf")


def parse_file(file_type: SourceFileType, content: bytes) -> ParseResult:
    if file_type == SourceFileType.CSV:
        return parse_csv(content)
    if file_type == SourceFileType.EXCEL:
        return parse_excel(content)
    if file_type == SourceFileType.PDF:
        return parse_pdf(content)
    raise ValueError(f"Unsupported file type: {file_type}")


def _rows_from_grid(grid: list[list[str | None]], warnings: list[str]) -> tuple[list[str], list[ParsedRow]]:
    """
    Shared header-detection heuristic: the first row with at least two
    non-empty cells is the header row; every non-blank row after it is
    data. Blank rows are skipped (common in supplier files as visual
    section breaks) but not silently dropped from row numbering awareness -
    row_number always reflects the row's ordinal position among the actual
    data rows kept, not its position in the source file (that raw source
    position isn't preserved here - a limitation worth knowing about, not
    hidden away).
    """
    header_idx = None
    for i, row in enumerate(grid):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_empty) >= 2:
            header_idx = i
            break
    if header_idx is None:
        warnings.append("Could not find a header row (no row with 2+ non-empty cells) - file may be empty")
        return [], []

    raw_headers = [str(c).strip() if c is not None else "" for c in grid[header_idx]]
    # De-duplicate blank/repeated header cells so dict keys stay stable and unique.
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        name = h or f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    rows: list[ParsedRow] = []
    row_number = 0
    for raw_row in grid[header_idx + 1 :]:
        cells = [str(c).strip() if c is not None else "" for c in raw_row]
        if not any(cells):
            continue  # blank visual-separator row, not data
        cells = (cells + [""] * len(headers))[: len(headers)]  # pad/truncate to header width
        row_number += 1
        rows.append(ParsedRow(row_number=row_number, raw_data=dict(zip(headers, cells, strict=True))))

    return headers, rows


def parse_csv(content: bytes) -> ParseResult:
    warnings: list[str] = []
    text = content.decode("utf-8-sig", errors="replace")  # utf-8-sig strips a BOM if Excel added one
    reader = csv.reader(io.StringIO(text))
    grid = list(reader)
    headers, rows = _rows_from_grid(grid, warnings)
    return ParseResult(headers=headers, rows=rows, warnings=warnings)


def parse_excel(content: bytes) -> ParseResult:
    warnings: list[str] = []
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if len(workbook.sheetnames) > 1:
        warnings.append(
            f"Workbook has {len(workbook.sheetnames)} sheets "
            f"({', '.join(workbook.sheetnames)}) - only the first (active) sheet was parsed"
        )
    sheet = workbook[workbook.sheetnames[0]]
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    headers, rows = _rows_from_grid(grid, warnings)
    return ParseResult(headers=headers, rows=rows, warnings=warnings)


def parse_pdf(content: bytes) -> ParseResult:
    warnings: list[str] = []
    combined_grid: list[list[str | None]] = []
    header_row: list[str | None] | None = None

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        if len(pdf.pages) == 0:
            warnings.append("PDF has no pages")
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if not tables:
                warnings.append(f"Page {page_num}: no table detected by pdfplumber")
                continue
            for table in tables:
                if not table:
                    continue
                if header_row is None:
                    header_row = table[0]
                    combined_grid.append(table[0])
                    combined_grid.extend(table[1:])
                elif table[0] == header_row:
                    # Repeated header on a later page (common in multi-page
                    # catalogs) - skip the duplicate, keep the data rows.
                    combined_grid.extend(table[1:])
                else:
                    combined_grid.extend(table)

    if header_row is None:
        warnings.append(
            "No tables detected anywhere in this PDF - it may be a scanned image (needs OCR, not "
            "supported yet) or a non-tabular marketing layout pdfplumber's default detection can't "
            "read. Route this file to manual review rather than trusting an empty result."
        )

    headers, rows = _rows_from_grid(combined_grid, warnings)
    return ParseResult(headers=headers, rows=rows, warnings=warnings)
